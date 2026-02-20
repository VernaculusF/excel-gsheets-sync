"""
Модуль для настройки стилей и форматирования Excel.
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """Класс для управления стилями Excel."""
    
    # Стиль заголовков
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    # Стиль данных
    TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    
    # Границы
    THIN_BORDER_SIDE = Side(border_style="thin", color="000000")
    BORDER = Border(
        left=THIN_BORDER_SIDE,
        right=THIN_BORDER_SIDE,
        top=THIN_BORDER_SIDE,
        bottom=THIN_BORDER_SIDE
    )
    
    # Условное форматирование (для значений > 100)
    HIGHLIGHT_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    HIGHLIGHT_THRESHOLD = 100
    
    @staticmethod
    def apply_header_style(ws: Worksheet, row: int = 1) -> None:
        """
        Применить стиль к строке заголовков.
        
        Args:
            ws: Рабочий лист Excel.
            row: Номер строки с заголовками (по умолчанию 1).
        """
        for cell in ws[row]:
            cell.font = ExcelFormatter.HEADER_FONT
            cell.fill = ExcelFormatter.HEADER_FILL
            cell.alignment = ExcelFormatter.HEADER_ALIGNMENT
            cell.border = ExcelFormatter.BORDER
        
        logger.debug(f"Применен стиль заголовков к строке {row}")
    
    @staticmethod
    def apply_data_style(ws: Worksheet, start_row: int = 2) -> None:
        """
        Применить стиль к ячейкам с данными.
        
        Args:
            ws: Рабочий лист Excel.
            start_row: Номер строки, с которой начинаются данные (по умолчанию 2).
        """
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                # Применяем границы ко всем ячейкам
                cell.border = ExcelFormatter.BORDER
                
                # Определяем выравнивание на основе типа значения
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = ExcelFormatter.NUMBER_ALIGNMENT
                    else:
                        cell.alignment = ExcelFormatter.TEXT_ALIGNMENT
        
        logger.debug(f"Применен стиль данных начиная со строки {start_row}")
    
    @staticmethod
    def apply_conditional_formatting(ws: Worksheet, start_row: int = 2) -> None:
        """
        Применить условное форматирование (значения > 100 выделяются зеленым).
        
        Args:
            ws: Рабочий лист Excel.
            start_row: Номер строки, с которой начинаются данные (по умолчанию 2).
        """
        highlighted_cells = 0
        
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    # Пытаемся преобразовать значение в число
                    try:
                        # Проверяем, что значение можно преобразовать в float
                        if isinstance(cell.value, (int, float)):
                            numeric_value = float(cell.value)
                        else:
                            numeric_value = float(str(cell.value))
                        
                        if numeric_value > ExcelFormatter.HIGHLIGHT_THRESHOLD:
                            cell.fill = ExcelFormatter.HIGHLIGHT_FILL
                            highlighted_cells += 1
                    except (ValueError, TypeError):
                        # Если значение не число, пропускаем
                        pass
        
        logger.debug(f"Условное форматирование: выделено {highlighted_cells} ячеек")
    
    @staticmethod
    def auto_adjust_column_width(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
        """
        Автоматически настроить ширину колонок по содержимому.
        
        Args:
            ws: Рабочий лист Excel.
            min_width: Минимальная ширина колонки.
            max_width: Максимальная ширина колонки.
        """
        for column in ws.columns:
            max_length = 0
            col_idx = column[0].column
            if col_idx is None:
                continue
            column_letter = get_column_letter(col_idx)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            
            # Устанавливаем ширину с учетом границ
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.debug("Автоматически настроена ширина колонок")
    
    @staticmethod
    def format_worksheet(ws: Worksheet) -> None:
        """
        Применить все стили форматирования к листу.
        
        Args:
            ws: Рабочий лист Excel.
        """
        logger.info("Применение форматирования к листу Excel...")
        
        # Применяем стили
        ExcelFormatter.apply_header_style(ws, row=1)
        ExcelFormatter.apply_data_style(ws, start_row=2)
        ExcelFormatter.apply_conditional_formatting(ws, start_row=2)
        ExcelFormatter.auto_adjust_column_width(ws)
        
        logger.info("✓ Форматирование успешно применено")
