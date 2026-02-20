"""
Модуль для работы с Excel файлами (чтение, запись, форматирование).
"""
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Any
from pathlib import Path
import logging
from formatter import ExcelFormatter

logger = logging.getLogger(__name__)


class ExcelHandler:
    """Класс для работы с Excel файлами."""
    
    def __init__(self):
        """Инициализация обработчика Excel."""
        self.formatter = ExcelFormatter()
    
    def create_excel(self, data: List[List[Any]], output_path: Path) -> None:
        """
        Создать Excel файл с данными и форматированием.
        
        Args:
            data: Данные для записи (список списков).
            output_path: Путь к выходному файлу.
            
        Raises:
            ValueError: Если данные пустые.
            IOError: Ошибка при записи файла.
        """
        if not data:
            raise ValueError("Данные для экспорта пустые")
        
        logger.info(f"Создание Excel файла {output_path}...")
        
        try:
            # Создаем новую рабочую книгу
            wb = Workbook()
            ws = wb.active
            assert ws is not None, "Не удалось получить активный лист"
            ws.title = "Data"
            
            # Записываем данные
            logger.info(f"Запись {len(data)} строк в Excel...")
            for row_idx, row_data in enumerate(data, start=1):
                # Заменяем None на пустую строку
                cleaned_row = ['' if cell is None else cell for cell in row_data]
                ws.append(cleaned_row)
            
            logger.info("✓ Данные записаны")
            
            # Применяем форматирование
            self.formatter.format_worksheet(ws)
            
            # Сохраняем файл
            wb.save(output_path)
            logger.info(f"✓ Файл сохранен: {output_path}")
            
        except IOError as e:
            logger.error(f"Ошибка при записи файла {output_path}: {e}")
            raise
        except Exception as e:
            logger.error(f"Неожиданная ошибка при создании Excel файла: {e}")
            raise
    
    def read_excel(self, file_path: Path) -> List[List[Any]]:
        """
        Прочитать данные из Excel файла.
        
        Args:
            file_path: Путь к Excel файлу.
            
        Returns:
            Список списков со значениями ячеек. Пустые ячейки представлены как ''.
            
        Raises:
            FileNotFoundError: Файл не найден.
            InvalidFileException: Неверный формат файла.
        """
        if not file_path.exists():
            logger.error(f"Файл {file_path} не найден")
            raise FileNotFoundError(f"Файл {file_path} не найден")
        
        logger.info(f"Чтение Excel файла {file_path}...")
        
        try:
            # Загружаем рабочую книгу (только данные, без форматирования)
            wb = load_workbook(file_path, data_only=True)
            
            # Берем первый лист
            ws = wb.active
            assert ws is not None, "Не удалось получить активный лист"
            logger.info(f"Используется лист: {ws.title}")
            
            # Читаем все значения
            data = []
            for row in ws.iter_rows(values_only=True):
                # Заменяем None на пустую строку
                cleaned_row = ['' if cell is None else cell for cell in row]
                data.append(cleaned_row)
            
            logger.info(f"✓ Прочитано {len(data)} строк")
            return data
            
        except InvalidFileException:
            logger.error(f"Файл {file_path} имеет неверный формат. Ожидается .xlsx")
            raise
        except Exception as e:
            logger.error(f"Ошибка при чтении файла {file_path}: {e}")
            raise
    
    def validate_structure(
        self, 
        excel_data: List[List[Any]], 
        expected_headers: List[str]
    ) -> tuple[bool, str]:
        """
        Проверить, что структура Excel файла соответствует ожидаемой.
        
        Args:
            excel_data: Данные из Excel файла.
            expected_headers: Ожидаемые заголовки из Google Sheets.
            
        Returns:
            Кортеж (успех, сообщение об ошибке).
        """
        if not excel_data:
            return False, "Excel файл пустой"
        
        excel_headers = excel_data[0]
        expected_col_count = len(expected_headers)
        actual_col_count = len(excel_headers)
        
        if actual_col_count != expected_col_count:
            error_msg = (
                f"Несоответствие структуры: "
                f"ожидается {expected_col_count} колонок, "
                f"но в Excel файле {actual_col_count} колонок.\n"
                f"Ожидаемые заголовки: {expected_headers}\n"
                f"Фактические заголовки: {excel_headers}"
            )
            logger.error(error_msg)
            return False, error_msg
        
        logger.info("✓ Структура Excel файла совпадает с Google Sheets")
        return True, ""
    
    def get_row_count(self, file_path: Path) -> int:
        """
        Получить количество строк в Excel файле.
        
        Args:
            file_path: Путь к Excel файлу.
            
        Returns:
            Количество строк.
            
        Raises:
            FileNotFoundError: Файл не найден.
        """
        data = self.read_excel(file_path)
        return len(data)
    
    def __repr__(self) -> str:
        """Строковое представление обработчика."""
        return "ExcelHandler()"
